# coding: utf-8

import logging
import webapp2
from ucf.utils.helpers import *
from ucf.pages.operator import *
import sateraito_inc
import sateraito_func
from ucf.utils.models import AnswerUser, ExcelTemplateFile
import urllib2
from openpyxl import load_workbook
import StringIO
from openpyxl.writer.excel import save_virtual_workbook
import lineworks_func
from google.appengine.api import namespace_manager
import json
from google.appengine.ext import blobstore
import cloudstorage
from google.appengine.api import app_identity
import string
import random
import urllib
from google.appengine.api import urlfetch


class Page(TenantAppHelper):

    def get(self):
        # try:
        # type = self.request.get("type").strip()
        # if type == 'pdf':
        # 	url = 'https://788260b5cb47.ngrok.io/?url='+base64.b64decode(self.request.get("url"))
        # 	response = urllib2.urlopen(url)
        # 	html = response.read()
        # 	response = urllib2.urlopen(html)
        # 	html = response.read()

        # 	self.response.headers['content-type'] = 'application/pdf'
        # 	self.response.headers['Content-Disposition'] = 'attachment; filename=file.pdf'
        # 	self.response.out.write(html)
        # 	return 1

        tenant = self.request.get("tenant")
        self._tenant = tenant
        namespace_manager.set_namespace(tenant.lower())

        session = self.request.get("session")

        # get answer
        q = AnswerUser.query()
        q = q.filter(AnswerUser.unique_id == session)

        answer = None

        for entry in q.iter(limit=20, offset=0):
            vo = entry.exchangeVo('Asia/Tokyo')
            list_vo = {}
            for k, v in vo.iteritems():
                if k in ['excel_blob', 'file_id', 'lineworks_id', 'pdf_blob', 'rule_id', 'value', 'sheet', 'unique_id']:
                    list_vo[k] = v
            answer = list_vo

        # get file
        q = ExcelTemplateFile.query()
        q = q.filter(ExcelTemplateFile.unique_id == answer['file_id'])
        file = None

        for entry in q.iter(limit=20, offset=0):
            vo = entry.exchangeVo('Asia/Tokyo')
            list_vo = {}
            for k, v in vo.iteritems():
                if k in ['blob_store']:
                    list_vo[k] = v
            file = list_vo

        data = urllib2.urlopen(sateraito_inc.my_site_url + "/tenant/template/download?blob_key=" + file['blob_store'])
        xlsx = data.read()
        wb = load_workbook(StringIO.StringIO(xlsx))
        wb.active = int(answer['sheet'])
        ws = wb.active
        val = json.loads(answer['value'])
        for item in val:
            ques = lineworks_func.findQuestionByAlias(item)
            if ques['location'].strip() != '' and val[item] == self.getMsg('SKIP') and ques['default'].strip() != '':
                ws[ques['location']] = ques['default']
            elif ques['location'].strip() != '':
                ws[ques['location']] = val[item]

        n = wb.sheetnames
        for i, val in enumerate(n):
            if i != int(answer['sheet']):
                del wb[val]

        bucket = app_identity.get_default_gcs_bucket_name()
        filename = '/{0}/{1}'.format(bucket,
                                     ''.join(random.choice(string.ascii_uppercase + string.digits) for _ in range(30)))
        with cloudstorage.open(filename, 'w') as filehandle:
            filehandle.write(save_virtual_workbook(wb))

        # /gs/bucket/object
        blobstore_filename_excel = '/gs{}'.format(filename)
        blob_key_excel = blobstore.BlobKey(blobstore.create_gs_key(blobstore_filename_excel))

        # convert excel to pdf via other service
        form_fields = {}
        form_fields['api_key'] = "sateraito_pdffaade815c82b459d9dbfd00e21a6d0cc"
        form_fields['file'] = sateraito_inc.my_site_url + "/tenant/template/download_cloudstorage/" + str(
            blob_key_excel) + "/xlsx"
        form_fields['file_type'] = "url"
        form_fields['nonce'] = '324534534543'

        form_data = urllib.urlencode(form_fields)
        request_url = "https://mlapi-dev.sateraito.jp/api/pdf/convert"

        deadline = 600
        headers = {
            'Accept': 'application/json;odata=verbose',
        }
        result = urlfetch.fetch(url=request_url, method="POST", payload=form_data, follow_redirects=True,
                                deadline=deadline, headers=headers)
        res = json.loads(result.content);
        data = urllib2.urlopen(res['file_url'])
        pdf = data.read()

        filename_pdf = '/{0}/{1}'.format(bucket, ''.join(
            random.choice(string.ascii_uppercase + string.digits) for _ in range(30)))
        with cloudstorage.open(filename_pdf, 'w') as filehandle:
            filehandle.write(pdf)

        # /gs/bucket/object
        blobstore_filename_pdf = '/gs{}'.format(filename_pdf)
        blob_key_pdf = blobstore.BlobKey(blobstore.create_gs_key(blobstore_filename_pdf))

        AnswerUser.update(answer['unique_id'], blob_key_pdf, blob_key_excel)

        self.response.headers['Content-Type'] = 'application/json'
        obj = {
            'status': True,
            'excel': str(blob_key_excel),
            'pdf': str(blob_key_pdf)
        }
        self.response.out.write(json.dumps(obj))
# except BaseException as e:
# 	print(e)
# 	self.response.out.write('<h1 style="text-align: center;">Something went wrong !!!</h1>')


app = ndb.toplevel(webapp2.WSGIApplication([('/tenant/template/download_excel', Page)], debug=sateraito_inc.debug_mode,
                                           config=sateraito_func.wsgi_config))
