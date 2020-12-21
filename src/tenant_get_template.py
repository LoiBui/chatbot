# coding: utf-8

import webapp2
from ucf.utils.helpers import *
from ucf.utils import loginfunc
from ucf.pages.operator import *
import sateraito_inc
import sateraito_func
from openpyxl import load_workbook
from io import BytesIO

_gnaviid = 'DASHBOARD'
_leftmenuid = 'INDEX'
class Page(TenantAppHelper):

	def post(self, tenant):
    		
		try:
			file = self.request.get('file')
			if not file:
				wb = load_workbook('excel/test.xlsx')
			else:
				wb = load_workbook(BytesIO(self.request.get('file')))

			jsondata = {
				'status': True,
				'data': wb.sheetnames
			}
			jsondata_str = json.JSONEncoder().encode(jsondata)
			self.response.out.write(jsondata_str)
		except BaseException as e:
			jsondata = {
				'status': False
			}
			jsondata_str = json.JSONEncoder().encode(jsondata)
			self.response.out.write(jsondata_str)
			pass

app = ndb.toplevel(webapp2.WSGIApplication([('/a/([^/]*)/get_template', Page)], debug=sateraito_inc.debug_mode, config=sateraito_func.wsgi_config))